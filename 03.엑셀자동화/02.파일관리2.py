import openpyxl

# 새로운 엑셀 파일 생성
wb = openpyxl.Workbook()

# 새로운 sheet 생성
ws = wb.create_sheet('2023.01')

# 모든 시트 이름 출력
print(wb.sheetnames) # ['Sheet', '2023.01']

# 시트삭제
del wb['Sheet']

# 엑셀저장
wb.save('03.엑셀자동화/거래처A매입현황.xlsx')